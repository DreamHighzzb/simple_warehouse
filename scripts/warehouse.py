from tkinter import *
import tkinter as tk
from tkinter import messagebox
import time
import xlsxwriter
from datetime import datetime
import scripts.combobox
import scripts.dropdownmenu
from openpyxl import load_workbook
import os
import math

root = None

#excel 写对象
global_wb = None
global_file_name = '仓库总表.xlsx'
global_model_file_name = '商品型号.xlsx'

# 示例数据 - 商品信息
inventory_data = ['商品型号','库存数量','入库数量','出库数量']

# 示例数据 - 入库记录
in_data = ['商品型号','入库数量','入库单价','供应商名','入库日期']

# 示例数据 - 出库记录
out_data = ['商品型号','出库数量','出库单价','经销商名','出库日期']

#商品规格
t_products = {}
t_products["商品型号"] = []

#展示frame
show_frame_main = None
show_frame_sub = None
show_frame_data = None

main_frame_title = {}
main_frame_title[1] = "入库操作"
main_frame_title[2] = "出库操作"

sheet_title = {}
sheet_title[1] = "入库表"
sheet_title[2] = "出库表"
sheet_title[3] = "库存表"

#展示数量
show_default_index = 1  #起始位置
show_gap_index = 10      #间隔
show_gap_index_ex = 16      #间隔
show_begin_index = show_default_index #数据表第二行为
filter_field_param = None
global_select_data = None #当前选择数据

#密码
password = "123987"

#下拉菜单
combo_box = []

def setPosition(e, posx, posy):
    e.place(x=posx, y=posy)

def setPositionAligned(e, posx, posy, ach="center"):
    e.place(x=posx, y=posy)
    e.place(relx=0.5, rely=0.5, anchor=ach)

def saveData(type):
    item = []
    for cbx in combo_box:
        item.append(cbx.entry.get())
        cbx.clear()
    if "" == item[0] or "" == item[1]:
        messagebox.askyesno("确认", "没输入商品型号或者入库数量")
        result = False
    else:
        result = messagebox.askyesno("确认", "商品型号: {0}\n入库数量: {1}\n入库单价: {2}\n供应商名: {3}\n".format(item[0], item[1], item[2], item[3]))
    if result:
        add_sheet_cell(sheet_title[type], item)

def filterFunc(filter_param, sheet_name):
    global show_frame_sub
    global filter_field_param
    filter_field_param = filter_param
    global show_begin_index
    show_begin_index = show_default_index
    global global_select_data
    global_select_data = None
    showTableData(sheet_name, show_frame_sub)

def showTableData(sheet_name,parent_frame, create_title = False):
    global global_wb
    if None == global_wb:
        return
    ws = global_wb[sheet_name]
    max_row = ws.max_row
    global show_begin_index
    beginIndex = show_begin_index
    endIndex = beginIndex + show_gap_index
    global show_frame_data
    if None != show_frame_data:
        show_frame_data.destroy()
    show_frame_data = tk.Frame(parent_frame,width=812, height=335)
    setPosition(show_frame_data, 0, -160)
    show_frame_data.lower()

    fieldList = ['商品型号','供应商名','入库日期']
    if sheet_name == sheet_title[2]:
        fieldList = ['商品型号','经销商名','出库日期']
    filterField = {}
    for field in fieldList:
        filterField[field] = []
        filterField[field].append(None)
    for i in range(1, ws.max_row + 1):
        if 1 != i:
            if ws[i][0].value not in filterField[fieldList[0]]:
                filterField[fieldList[0]].append(ws[i][0].value)
            if ws[i][3].value not in filterField[fieldList[1]]:
                filterField[fieldList[1]].append(ws[i][3].value)
            if ws[i][4].value not in filterField[fieldList[2]]:
                filterField[fieldList[2]].append(ws[i][4].value)
    rowIdx = 1
    global global_select_data
    if None == global_select_data:
        global_select_data = {}
        for i in range(2, ws.max_row + 1):
            if None == filter_field_param or filter_field_param in [ws[i][0].value, ws[i][3].value,ws[i][4].value]:
                global_select_data[rowIdx] = []
                global_select_data[rowIdx].append(i)
                global_select_data[rowIdx].append(ws[i][0].value)
                global_select_data[rowIdx].append(ws[i][1].value)
                global_select_data[rowIdx].append(ws[i][2].value)
                global_select_data[rowIdx].append(ws[i][3].value)
                global_select_data[rowIdx].append(ws[i][4].value)
                rowIdx = rowIdx + 1

    gapIndex = 0
    for i, v in global_select_data.items():
        if beginIndex <= i and i < endIndex:
            gapIndex = gapIndex + 1
            colPos = -175 + 32 * gapIndex
            lb = tk.Label(show_frame_data, text=v[0],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, -392, colPos, "center")
            lb = tk.Label(show_frame_data, text=v[1],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, -295, colPos, "center")
            lb = tk.Label(show_frame_data, text=v[2],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, -150, colPos, "center")
            lb = tk.Label(show_frame_data, text=v[3],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, 0, colPos, "center")
            lb = tk.Label(show_frame_data, text=v[4],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, 150, colPos, "center")
            lb = tk.Label(show_frame_data, text=v[5],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, 305, colPos, "center")
            delBtn = tk.Button(
            show_frame_data,
            text="X",
            width=2,
            height=1,
            command=lambda idx=v[0]:deleteButtonFunc(deleteShowData, sheet_name, idx),
            bg='#2196F3', 
            fg='white',
            font=("Arial", 10)
            )
            setPositionAligned(delBtn, 390, colPos, "center")

    if create_title:
        for cell in ws:
            colIdx = 0
            lb = tk.Label(parent_frame, text="编号",font=("Arial", 12, "bold"))
            setPosition(lb, -5, -185)
            for item in cell:
                if item.value in filterField:
                    cbox = scripts.dropdownmenu.createFilteredMenu(parent_frame, filterField[item.value], 20, 3, item.value, filterFunc, sheet_name) 
                    setPosition(cbox, 70 + colIdx * 150, -185)
                else:
                    lb = tk.Label(parent_frame, text=item.value,font=("Arial", 12, "bold"))
                    setPosition(lb, 70 + colIdx * 150, -185)
                colIdx = colIdx + 1
            break      

def showTableDataEx(sheet_name,parent_frame, saveData):
    global global_wb
    if None == global_wb:
        return

    ws = global_wb[sheet_title[3]]
    max_row = ws.max_row
    global show_begin_index
    beginIndex = show_begin_index
    endIndex = beginIndex + show_gap_index_ex
    global show_frame_data
    if None != show_frame_data:
        show_frame_data.destroy()
    show_frame_data = tk.Frame(parent_frame,width=812, height=528)
    setPosition(show_frame_data, 0, -270)
    in_ws = global_wb[sheet_title[1]]
    out_ws = global_wb[sheet_title[2]]
    global global_select_data
    if None == global_select_data:
        global_select_data = {}
        index = 0
        for cell in in_ws:
            if index > 0:
                key = cell[0].value
                num = cell[1].value
                if not global_select_data.get(key):
                    global_select_data[key] = [0, 0, 0]
                global_select_data[key][0] = global_select_data[key][0] + int(num)
                global_select_data[key][1] = global_select_data[key][1] + int(num)
            index = index + 1

        index = 0
        for cell in out_ws:
            if index > 0:
                key = cell[0].value
                num = cell[1].value
                if not global_select_data.get(key):
                    global_select_data[key] = [0, 0, 0]
                global_select_data[key][2] = global_select_data[key][2] + int(num)
                global_select_data[key][0] = global_select_data[key][0] - int(num)
            index = index + 1

    colIdx = 0
    lb = tk.Label(show_frame_data, text='编号',font=("Arial", 12, "bold"))
    setPosition(lb, 0, 10)
    for field in inventory_data:
        lb = tk.Label(show_frame_data, text=field,font=("Arial", 12, "bold"))
        setPosition(lb, 90 + colIdx * 145, 10)
        colIdx = colIdx + 1

    countIdx = 0
    saveRowIdx = 2
    realIdx = 0
    for k in global_select_data:
        countIdx = countIdx + 1
        if countIdx > endIndex and False == saveData:
            break
        if beginIndex <= countIdx and countIdx < endIndex:
            realIdx = realIdx + 1
            colPos = -250 + 30 * realIdx
            lb = tk.Label(show_frame_data, text=countIdx,font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, -390, colPos, "center")
            lb = tk.Label(show_frame_data, text=k,font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, -280, colPos, "center")
            lb = tk.Label(show_frame_data, text=global_select_data[k][0],font=("黑体", 16),fg='red',anchor="center")
            setPositionAligned(lb, -140, colPos, "center")
            lb = tk.Label(show_frame_data, text=global_select_data[k][1],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, 5, colPos, "center")
            lb = tk.Label(show_frame_data, text=global_select_data[k][2],font=("黑体", 16),fg='blue',anchor="center")
            setPositionAligned(lb, 150, colPos, "center")
        if saveData and ws:
            ws.cell(row=saveRowIdx, column = 1, value=k)
            ws.cell(row=saveRowIdx, column = 2, value=global_select_data[k][0])
            ws.cell(row=saveRowIdx, column = 3, value=global_select_data[k][1])
            ws.cell(row=saveRowIdx, column = 4, value=global_select_data[k][2])
            saveRowIdx = saveRowIdx + 1

    if saveData and ws:
        global_wb.save(global_file_name)
    
def saveButtonFunc(event_func, type):
    event_func(type)

def pageShowFunc(event_func, type):
    event_func(type)

def deleteShowData(sheet_name, index):
    result = messagebox.askyesno("确认", "确认删除数据？")
    if result:
        remove_sheet_cell(sheet_name, index)

def deleteButtonFunc(event_func, sheet_name, index):
    event_func(sheet_name, index)

def preShow(type):
    global show_begin_index
    global sheet_title
    global show_frame_sub
    if show_default_index == show_begin_index:
            return
    if 3 != type:
        show_begin_index = show_begin_index - show_gap_index
        if show_begin_index < show_default_index:
            show_begin_index = show_default_index
        showTableData(sheet_title[type],show_frame_sub)
    else:
        if show_default_index == show_begin_index:
            return
        show_begin_index = show_begin_index - show_gap_index_ex
        if show_begin_index < show_default_index:
            show_begin_index = show_default_index
        showTableDataEx(sheet_title[type],show_frame_sub, False)

def nextShow(type):
    global show_begin_index
    global sheet_title
    global show_frame_sub
    global global_select_data
    max_index = 0
    if None != global_select_data:
        max_index = len(global_select_data)
    if 3 != type:
        show_begin_index = show_begin_index + show_gap_index
        if show_begin_index > max_index:
            show_begin_index = show_begin_index - show_gap_index
            return
        showTableData(sheet_title[type],show_frame_sub)
    else:
        show_begin_index = show_begin_index + show_gap_index_ex
        if show_begin_index > max_index:
            show_begin_index = show_begin_index - show_gap_index_ex
            return
        showTableDataEx(sheet_title[type],show_frame_sub, False)

def showView(type):
    global show_frame_main
    if None != show_frame_main:
        show_frame_main.destroy()

    global show_frame_sub
    if None != show_frame_sub:
        show_frame_sub.destroy()

    global filter_field_param
    filter_field_param = None

    global combo_box
    for cbox in combo_box:
        del cbox
    combo_box = []
    global show_begin_index
    show_begin_index = show_default_index

    global global_select_data
    global_select_data = None

    if 3 != type:
        show_frame_main = tk.LabelFrame(root, text=main_frame_title[type], font=("Arial", 13, "bold"), bg="#f0f0f0", padx=10, pady=50)
        setPosition(show_frame_main, 10, 0)
        title_label = tk.Label(show_frame_main, text="", font=("Arial", 20, "bold"), bg="#f0f0f0", fg="#333")
        title_label.pack(padx = 400, pady=0)
        tag = 0
        field_data = None
        if 1 == type:
            field_data = in_data
        elif 2 == type:
            field_data = out_data
        for field in field_data:
            lb = tk.Label(show_frame_main, text=field,font=("Arial", 12, "bold"))
            if field in ['商品型号']:
                cbox = scripts.combobox.createComboBox(show_frame_main, t_products[field], False, True, False, 20) 
                setPosition(lb, 45 + tag * 150, -50)
                setPosition(cbox, tag * 150, -25)
            else:
                cbox = scripts.combobox.createComboBox(show_frame_main, "",False, False, False,10) 
                if field in ['入库日期','出库日期']:
                    cbox.is_date_lab = True
                    cbox._cancel_bind_events()
                setPosition(lb, 95 + tag * 110, -50)
                setPosition(cbox, 90 + tag * 110, -25)
            combo_box.append(cbox)
            tag = tag + 1
        button = tk.Button(show_frame_main, text="保存", command=lambda:saveButtonFunc(saveData, type),
            font=('微软雅黑', 12),bg='#2196F3', fg='white',activebackground='#45a049',padx=5,pady=5,relief=tk.FLAT)
        setPosition(button, tag * 150 + 8, 40)

        show_frame_sub = tk.LabelFrame(root, text=sheet_title[type], font=("Arial", 13, "bold"), bg="#f0f0f0", padx=10, pady=186)
        setPosition(show_frame_sub, 10, 165)
        title_label = tk.Label(show_frame_sub, text="", font=("Arial", 20, "bold"), bg="#f0f0f0")
        title_label.pack(padx = 400, pady=0)
        preBtn = tk.Button(
            show_frame_sub,
            text="<",
            width=3,
            command=lambda:pageShowFunc(preShow, type),
            bg='#2196F3', 
            fg='white',
            font=("微软雅黑", 14)
        )
        setPosition(preBtn, 320, 175)
        nextBtn = tk.Button(
            show_frame_sub,
            text=">",
            width=3,
            command=lambda:pageShowFunc(nextShow, type),
            bg='#2196F3', 
            fg='white',
            font=("微软雅黑", 14)
        )
        setPosition(nextBtn, 420, 175)
        showTableData(sheet_title[type],show_frame_sub, True)
    else:
        show_frame_sub = tk.LabelFrame(root, text=sheet_title[type], font=("Arial", 13, "bold"), bg="#f0f0f0", padx=10, pady=269)
        setPosition(show_frame_sub, 10, -1)
        title_label = tk.Label(show_frame_sub, text="", font=("Arial", 20, "bold"), bg="#f0f0f0")
        title_label.pack(padx = 400, pady=0)
        preBtn = tk.Button(
            show_frame_sub,
            text="<",
            width=3,
            command=lambda:pageShowFunc(preShow, type),
            bg='#2196F3', 
            fg='white',
            font=("微软雅黑", 14)
        )
        setPosition(preBtn, 320, 258)
        nextBtn = tk.Button(
            show_frame_sub,
            text=">",
            width=3,
            command=lambda:pageShowFunc(nextShow, type),
            bg='#2196F3', 
            fg='white',
            font=("微软雅黑", 14)
        )
        setPosition(nextBtn, 420, 258)
        showTableDataEx(sheet_title[type],show_frame_sub, True)

def protect_all_sheets(workbook, password="password"):
    """
    保护工作簿中的所有工作表
    """
    global global_wb
    for sheet_name in workbook.sheetnames:
        ws = global_wb[sheet_name]
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.enable()
        # # 设置保护选项（可选）
        # sheet.protection.autoFilter = False  # 不允许使用自动筛选
        # sheet.protection.deleteColumns = False  # 不允许删除列
        # sheet.protection.deleteRows = False  # 不允许删除行
        # sheet.protection.formatCells = False  # 不允许格式化单元格
        # sheet.protection.insertColumns = False  # 不允许插入列
        # sheet.protection.insertRows = False  # 不允许插入行
        # sheet.protection.objects = False  # 不允许修改对象
        # sheet.protection.scenarios = False  # 不允许修改方案
        # sheet.protection.selectLockedCells = False  # 不允许选择锁定单元格
        # sheet.protection.selectUnlockedCells = False  # 不允许选择未锁定单元格
        # sheet.protection.sort = False  # 不允许排序
    global_wb.save(global_file_name)

#创建总表
def createExcelTable():
    # 创建工作簿
    # filename = f"仓库总表_{datetime.now().strftime('%Y%m%d')}.xlsx"
    global global_file_name
    workbook = xlsxwriter.Workbook(global_file_name)

    # 设置全局格式
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': 'white',
        'bg_color': '#366092',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'border': 1
    })

    # 设置数字格式
    currency_format = workbook.add_format({
        'num_format': '#,##0.000',
        'border': 1
    })
    
    number_format = workbook.add_format({
        'num_format': '0',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd',
        'border': 1
    })
    
    datetime_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd hh:mm',
        'border': 1
    })
    
    normal_format = workbook.add_format({
        'border': 1
    })

    global sheet_title
    worksheet_in = workbook.add_worksheet(sheet_title[1])
    for col, header in enumerate(in_data):
        worksheet_in.write(0, col, header, header_format)

    worksheet_out = workbook.add_worksheet(sheet_title[2])
    for col, header in enumerate(out_data):
        worksheet_out.write(0, col, header, header_format)

    worksheet_inventory = workbook.add_worksheet(sheet_title[3])
    for col, header in enumerate(inventory_data):
        worksheet_inventory.write(0, col, header, header_format)

    # 关闭工作簿
    workbook.close()
    
    print(f"使用xlsxwriter生成的Excel文件: {global_file_name}")
    return global_file_name

def createMenuUI(root):
    root.title('财富仓库')
    root.geometry('{0}x{1}'.format(850, 600)) # 这里的乘号不是 * ，而是小写英文字母 x
    root.resizable(False, False)
    root.configure(bg="#f0f0f0")
    """创建菜单栏"""
    menubar = tk.Menu(root)
    root.config(menu=menubar)

    welcomeLab = tk.Label(
        root,
        text="欢迎使用财富仓库，选择菜单栏进入仓库!!",
        font=("微软雅黑", 20),
        fg='blue',
        width=50,
        height=30
    )
    welcomeLab.pack(expand=True, anchor="center")  # expand=True 填充可用空间
    
    # 文件菜单
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="菜单栏", menu=file_menu)
    file_menu.add_command(label="入库信息",command=lambda:showView(1))
    file_menu.add_command(label="出库信息",command=lambda:showView(2))
    file_menu.add_command(label="库存信息",command=lambda:showView(3))

def read_sheets_with_format(file_name,is_first):
    # 加载工作簿
    global global_wb
    global_wb = load_workbook(file_name)
    if is_first:
        protect_all_sheets(global_wb, password)
    model_wb = load_workbook(global_model_file_name)
    sheet_names = model_wb.sheetnames
    for i, sheet_name in enumerate(sheet_names[:3], 1):
        ws = model_wb[sheet_name]
        colIdx = 0
        for cell in ws:
            if colIdx > 0:
                t_products["商品型号"].append(cell[0].value)
            colIdx = colIdx + 1

    # print(f"文件中的表单: {sheet_names}")
    
    # # 分别处理前三个表单
    # for i, sheet_name in enumerate(sheet_names[:3], 1):
    #     ws = global_wb[sheet_name]
    #     show_sheet(ws)

def add_sheet_cell(sheet_name, item):
    global global_wb
    if None == global_wb:
        return
    ws = global_wb[sheet_name]
    last_row = ws.max_row + 1
    col = 1
    for v in item:
        ws.cell(row=last_row, column = col, value=v)
        col = col + 1
    global_wb.save(global_file_name)
    global show_frame_sub
    global global_select_data
    global_select_data = None
    showTableData(sheet_name, show_frame_sub)

def remove_sheet_cell(sheet_name, index):
    global global_wb
    if None == global_wb:
        return
    ws = global_wb[sheet_name]
    index = index
    if index <= 1 or index > ws.max_row:
        return
    ws.delete_rows(index, 1)  # 第一个参数是开始删除的行号，第二个参数是删除的行数
    global_wb.save(global_file_name)
    global show_frame_sub
    global global_select_data
    global_select_data = None
    showTableData(sheet_name, show_frame_sub)

def show_sheet(worksheet):
    for cell in worksheet:
        for item in cell:
            print(item.value)

def initDefaultUI(root):
    createMenuUI(root)

def run():
    isFirst = False
    if not os.path.exists(global_file_name):
        createExcelTable()
        isFirst = True
    read_sheets_with_format(global_file_name,isFirst)
    global root
    root = tk.Tk()
    initDefaultUI(root)
    root.mainloop()    
    
